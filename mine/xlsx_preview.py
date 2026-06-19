"""Build a minimal HTML table preview for .xlsx uploads (all sheets, bounded size per sheet)."""

from __future__ import annotations

import logging
from html import escape
from io import BytesIO

logger = logging.getLogger(__name__)

_MAX_ROWS = 120
_MAX_COLS = 28
_MAX_SHEETS = 30


def _slug_sheet_id(index: int) -> str:
    return f"sheet-{index}"


def _render_sheet_rows(ws, max_rows: int, max_cols: int) -> tuple[str, int]:
    """Return (tbody inner HTML, data row count)."""
    chunks: list[str] = []
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        row_count += 1
        chunks.append("<tr>")
        for cell in row:
            val = "" if cell is None else str(cell)
            chunks.append(f"<td>{escape(val)}</td>")
        chunks.append("</tr>")
    inner = "".join(chunks)
    if row_count == 0:
        inner = '<tr><td class="empty-sheet">No cells in this sheet.</td></tr>'
    return inner, row_count


def render_xlsx_preview_html(
    data: bytes,
    *,
    workbook_title: str = "",
    max_rows: int = _MAX_ROWS,
    max_cols: int = _MAX_COLS,
    max_sheets: int = _MAX_SHEETS,
) -> str | None:
    """
    Return a full HTML document with one bounded table per worksheet, or None if unreadable.

    workbook_title: optional file name shown in <title> / heading (not a sheet name).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not available; skipping xlsx HTML preview")
        return None

    if not data:
        return None

    try:
        wb = load_workbook(BytesIO(data), read_only=False, data_only=True)
    except Exception as e:
        logger.info("xlsx preview could not open workbook: %s", e)
        return None

    try:
        if not wb.worksheets:
            return None

        all_sheets = list(wb.worksheets)
        total = len(all_sheets)
        sheets = all_sheets[:max_sheets]
        truncated = total > max_sheets

        doc_title = escape((workbook_title or "Workbook").strip() or "Workbook")

        parts: list[str] = [
            "<!DOCTYPE html>",
            '<html lang="en"><head><meta charset="utf-8"/>',
            '<meta name="viewport" content="width=device-width, initial-scale=1"/>',
            f"<title>{doc_title} · preview</title>",
            "<style>",
            "body{margin:0;font-family:system-ui,-apple-system,Segoe UI,Roboto,sans-serif;font-size:12px;background:#f1f5f9;color:#0f172a;}",
            ".wrap{box-sizing:border-box;padding:10px 10px 24px;max-width:100%;}",
            "header.page-hd{margin-bottom:12px;}",
            "header.page-hd h1{margin:0 0 6px;font-size:1.05rem;font-weight:600;color:#0f172a;}",
            ".toc{background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:10px 12px;margin-bottom:16px;box-shadow:0 1px 2px rgba(15,23,42,.06);}",
            ".toc-label{display:block;font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.04em;margin-bottom:8px;}",
            ".toc ul{list-style:none;margin:0;padding:0;display:flex;flex-wrap:wrap;gap:6px 10px;}",
            ".toc a{color:#103f5a;font-weight:500;text-decoration:none;border-bottom:1px solid transparent;}",
            ".toc a:hover{border-bottom-color:#103f5a;}",
            ".sheet{margin-bottom:28px;scroll-margin-top:12px;}",
            ".sheet-hd{margin:0 0 8px;font-size:0.95rem;font-weight:600;color:#334155;display:flex;align-items:center;gap:8px;flex-wrap:wrap;}",
            ".sheet-index{font-size:11px;font-weight:500;color:#64748b;background:#e2e8f0;padding:2px 8px;border-radius:999px;}",
            ".table-scroll{overflow:auto;max-height:min(70vh,560px);border:1px solid #e2e8f0;border-radius:8px;background:#fff;}",
            "table{border-collapse:collapse;width:100%;min-width:max-content;}",
            "td{border:1px solid #e2e8f0;padding:5px 8px;white-space:nowrap;max-width:260px;overflow:hidden;text-overflow:ellipsis;vertical-align:top;}",
            "td.empty-sheet{border:none;color:#64748b;font-style:italic;padding:12px;}",
            ".note{font-size:11px;color:#64748b;margin:16px 0 0;line-height:1.45;}",
            ".warn{font-size:11px;color:#92400e;background:#fffbeb;border:1px solid #fcd34d;border-radius:6px;padding:8px 10px;margin-bottom:12px;}",
            "</style></head><body><div class=\"wrap\">",
            "<header class=\"page-hd\">",
            f"<h1>{doc_title}</h1>",
            f"<p class=\"note\" style=\"margin:0;\">Preview: each sheet shows up to {max_rows} rows × {max_cols} columns. Download the original for the full workbook.</p>",
            "</header>",
        ]

        if truncated:
            parts.append(
                f'<p class="warn">This file has <strong>{total}</strong> sheets; showing the first <strong>{max_sheets}</strong> only in this preview.</p>'
            )

        parts.append('<nav class="toc" aria-label="Jump to sheet">')
        parts.append('<span class="toc-label">Sheets in this workbook</span><ul>')
        for i, ws in enumerate(sheets):
            name = escape((ws.title or f"Sheet{i + 1}").strip() or f"Sheet{i + 1}")
            sid = _slug_sheet_id(i)
            parts.append(f'<li><a href="#{sid}">{name}</a></li>')
        parts.append("</ul></nav>")

        for i, ws in enumerate(sheets):
            sid = _slug_sheet_id(i)
            sheet_name = (ws.title or f"Sheet{i + 1}").strip() or f"Sheet{i + 1}"
            tbody_inner, _ = _render_sheet_rows(ws, max_rows, max_cols)
            parts.append(f'<section class="sheet" id="{sid}" aria-labelledby="{sid}-title">')
            parts.append(
                f'<h2 class="sheet-hd" id="{sid}-title"><span class="sheet-index">{i + 1} / {total}</span>'
                f"<span>{escape(sheet_name)}</span></h2>"
            )
            parts.append('<div class="table-scroll"><table><tbody>')
            parts.append(tbody_inner)
            parts.append("</tbody></table></div></section>")

        parts.append("</div></body></html>")
        return "".join(parts)
    finally:
        try:
            wb.close()
        except Exception:
            pass
