"""Parse Freeport active resource tracker Excel — Name + TSR for roster display and user import."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from flask import current_app
from werkzeug.datastructures import FileStorage

_NAME_HEADERS = frozenset(
    {
        "name",
        "resource name",
        "employee name",
        "full name",
        "consultant name",
        "associate name",
        "team member",
        "team member name",
    }
)
_TSR_HEADERS = frozenset(
    {
        "tsr",
        "tsr id",
        "tsr #",
        "tsr number",
        "tsr no",
        "tsr no.",
        "tsrid",
    }
)
_EMP_ID_HEADERS = frozenset(
    {
        "emp id",
        "empid",
        "emp_id",
        "employee id",
        "employeeid",
        "employee #",
        "employee no",
        "employee no.",
        "associate id",
        "associateid",
    }
)

ROSTER_FILENAME = "Freeport active resource Tracker - Updated.xlsx"

# Shared first-login password — users change it under Account settings.
DEFAULT_ROSTER_PASSWORD = "Mine123!"


@dataclass(frozen=True)
class TeamRosterRow:
    name: str
    tsr: str
    row_num: int
    emp_id: str = ""


@dataclass(frozen=True)
class ProposedUser:
    name: str
    tsr: str
    username: str
    email: str
    password: str
    row_num: int
    emp_id: str = ""


def roster_xlsx_path() -> Path:
    raw = (current_app.config.get("TEAM_ROSTER_XLSX") or "").strip()
    if raw:
        p = Path(raw)
        if not p.is_absolute():
            p = (Path(current_app.config["BASE_DIR"]) / p).resolve()
        return p
    return (Path(current_app.config["BASE_DIR"]) / "data" / ROSTER_FILENAME).resolve()


def save_roster_workbook(upload: FileStorage) -> Path:
    """Save uploaded tracker workbook to the configured data path."""
    dest = roster_xlsx_path()
    dest.parent.mkdir(parents=True, exist_ok=True)
    upload.save(dest)
    return dest


def _norm_header(cell) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell).strip().lower())


def _cell_str(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, float) and cell == int(cell):
        return str(int(cell))
    return str(cell).strip()


def _find_header_row(ws, *, max_scan: int = 25) -> tuple[int, int, int, int | None] | None:
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        name_col = tsr_col = emp_col = None
        for c_idx, cell in enumerate(row):
            h = _norm_header(cell)
            if not h:
                continue
            if h in _NAME_HEADERS or (name_col is None and h.endswith(" name")):
                name_col = c_idx
            if h in _TSR_HEADERS or h.replace(" ", "") == "tsr":
                tsr_col = c_idx
            if h in _EMP_ID_HEADERS or h.replace(" ", "") in ("empid", "employeeid"):
                emp_col = c_idx
        if name_col is not None and tsr_col is not None:
            return r_idx, name_col, tsr_col, emp_col
    return None


def load_team_roster(path: Path | None = None) -> list[TeamRosterRow]:
    path = path or roster_xlsx_path()
    if not path.is_file():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    rows: list[TeamRosterRow] = []
    try:
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            found = _find_header_row(ws)
            if not found:
                continue
            header_row, name_col, tsr_col, emp_col = found
            for r_idx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not row:
                    continue
                name = _cell_str(row[name_col]) if name_col < len(row) else ""
                tsr = _cell_str(row[tsr_col]) if tsr_col < len(row) else ""
                emp_id = ""
                if emp_col is not None and emp_col < len(row):
                    emp_id = _cell_str(row[emp_col])
                if not name and not tsr:
                    continue
                low = name.lower()
                if low in ("name", "resource name", "total", "grand total"):
                    continue
                if not name:
                    continue
                rows.append(TeamRosterRow(name=name, tsr=tsr, row_num=r_idx, emp_id=emp_id))
            if rows:
                break
    finally:
        wb.close()

    return rows


def roster_member_count(path: Path | None = None) -> int:
    return len(load_team_roster(path))


_DISCIPLINE_ORDER = (
    "Architecture",
    "Enterprise SAP",
    "Big Data & Analytics",
    "Data Integration",
    "Information Delivery",
    "PMO & Leadership",
    "Delivery",
)

_AVATAR_TONES = ("cyan", "teal", "copper", "azure", "slate")


def roster_discipline(tsr: str) -> str:
    """Map TSR title to a squad label for grouped display."""
    t = (tsr or "").lower()
    if "pmo" in t or "talent management" in t:
        return "PMO & Leadership"
    if "information delivery" in t or "visualization" in t:
        return "Information Delivery"
    if "big data" in t:
        return "Big Data & Analytics"
    if "data integration" in t or ("integration" in t and "architect" in t):
        return "Data Integration"
    if "integration" in t:
        return "Data Integration"
    if "sap" in t or t.startswith("es "):
        return "Enterprise SAP"
    if "architect" in t:
        return "Architecture"
    return "Delivery"


def roster_initials(name: str) -> str:
    parts = [p for p in re.sub(r"[^a-zA-Z\s]", "", name or "").split() if p]
    if len(parts) >= 2:
        return (parts[0][0] + parts[-1][0]).upper()
    if parts:
        return parts[0][:2].upper()
    return "?"


def roster_avatar_tone(name: str) -> str:
    h = sum(ord(c) for c in (name or ""))
    return _AVATAR_TONES[h % len(_AVATAR_TONES)]


def _discipline_slug(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")


def group_roster_for_display(roster: list[TeamRosterRow]) -> list[dict]:
    """Group roster rows by discipline with display metadata for the team page."""
    buckets: dict[str, list[dict]] = {}
    for row in roster:
        discipline = roster_discipline(row.tsr)
        buckets.setdefault(discipline, []).append(
            {
                "name": row.name,
                "tsr": row.tsr,
                "initials": roster_initials(row.name),
                "tone": roster_avatar_tone(row.name),
                "discipline": discipline,
            }
        )

    order = {label: idx for idx, label in enumerate(_DISCIPLINE_ORDER)}
    groups: list[dict] = []
    for label in sorted(buckets.keys(), key=lambda k: (order.get(k, 99), k)):
        slug = _discipline_slug(label)
        groups.append(
            {
                "label": label,
                "slug": slug,
                "count": len(buckets[label]),
                "members": buckets[label],
            }
        )
    return groups


def _slug_username(name: str) -> str:
    parts = re.sub(r"[^a-zA-Z0-9\s.-]", "", name).strip().lower().split()
    if len(parts) >= 2:
        base = f"{parts[0]}.{parts[-1]}"
    elif parts:
        base = parts[0]
    else:
        base = "user"
    return re.sub(r"[^a-z0-9.]", "", base)[:40] or "user"


def _username_from_emp_id(emp_id: str) -> str | None:
    raw = re.sub(r"[^a-zA-Z0-9._-]", "", (emp_id or "").strip())
    if len(raw) >= 3:
        return raw[:60]
    return None


def _username_from_tsr(tsr: str) -> str | None:
    raw = re.sub(r"[^a-zA-Z0-9._-]", "", (tsr or "").strip().lower())
    if len(raw) >= 3:
        return raw[:60]
    return None


def propose_users(
    roster: list[TeamRosterRow],
    *,
    existing_usernames: set[str] | None = None,
    existing_emails: set[str] | None = None,
    default_password: str = DEFAULT_ROSTER_PASSWORD,
) -> list[ProposedUser]:
    taken_usernames = {u.lower() for u in (existing_usernames or set())}
    taken_emails = {e.lower() for e in (existing_emails or set())}
    out: list[ProposedUser] = []
    password = (default_password or DEFAULT_ROSTER_PASSWORD).strip() or DEFAULT_ROSTER_PASSWORD

    for row in roster:
        base = (
            _username_from_emp_id(row.emp_id)
            or _username_from_tsr(row.tsr)
            or _slug_username(row.name)
        )
        username = base
        n = 2
        while username.lower() in taken_usernames:
            username = f"{base}{n}"
            n += 1
        taken_usernames.add(username.lower())

        email_base = username.replace(" ", ".")
        email = f"{email_base}@hexaware.local"
        if email.lower() in taken_emails:
            email = f"{email_base}+{row.row_num}@hexaware.local"
        taken_emails.add(email.lower())

        out.append(
            ProposedUser(
                name=row.name,
                tsr=row.tsr,
                username=username,
                email=email,
                password=password,
                row_num=row.row_num,
                emp_id=row.emp_id,
            )
        )
    return out


def import_roster_users(db, *, actor_id: int | None = None) -> dict:
    """Create MiNe users from roster; username = EMP ID, display name = NAME."""
    import bcrypt
    import csv

    from mine.services import log_audit

    roster = load_team_roster()
    if not roster:
        return {"created": 0, "skipped": 0, "error": "Roster file missing or no Name/TSR columns found."}

    missing_emp = sum(1 for r in roster if not (r.emp_id or "").strip())
    if missing_emp:
        return {
            "created": 0,
            "skipped": 0,
            "error": f"Roster has {missing_emp} row(s) without EMP ID. Fix the workbook and retry.",
        }

    existing = db.execute("SELECT username, email FROM users").fetchall()
    existing_usernames = {r["username"] for r in existing}
    existing_emails = {r["email"] for r in existing}

    proposed = propose_users(
        roster,
        existing_usernames=existing_usernames,
        existing_emails=existing_emails,
        default_password=DEFAULT_ROSTER_PASSWORD,
    )

    created_rows: list[ProposedUser] = []
    skipped = 0
    for pu in proposed:
        if pu.username.lower() in {u.lower() for u in existing_usernames} or pu.email.lower() in {
            e.lower() for e in existing_emails
        }:
            skipped += 1
            continue
        pw_hash = bcrypt.hashpw(pu.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        try:
            db.execute(
                """
                INSERT INTO users (username, email, password_hash, display_name, role, is_active)
                VALUES (?, ?, ?, ?, 'user', 1)
                """,
                (pu.username, pu.email.lower(), pw_hash, pu.name),
            )
            existing_usernames.add(pu.username)
            existing_emails.add(pu.email.lower())
            created_rows.append(pu)
            if actor_id is not None:
                log_audit(actor_id, "user_create", "user", None, pu.username)
        except Exception:
            skipped += 1

    cred_path = None
    if created_rows:
        log_dir = Path(current_app.config["BASE_DIR"]) / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        cred_path = log_dir / "team-roster-import.csv"
        with cred_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["name", "emp_id", "tsr", "username", "email", "default_password"])
            for pu in created_rows:
                writer.writerow([pu.name, pu.emp_id, pu.tsr, pu.username, pu.email, pu.password])

    return {
        "created": len(created_rows),
        "skipped": skipped,
        "default_password": DEFAULT_ROSTER_PASSWORD if created_rows else None,
        "credentials_file": str(cred_path) if cred_path else None,
        "error": None,
    }
